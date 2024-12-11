import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import logging
import re

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class StructuredNotesProcessor:
    def __init__(self, base_file, gbil_file, cash_file):
        self.base_file = base_file
        self.gbil_file = gbil_file
        self.cash_file = cash_file

    def validate_files(self):
        """Validate all input files before processing."""
        try:
            if not all([
                os.path.exists(self.base_file),
                os.path.exists(self.gbil_file),
                os.path.exists(self.cash_file)
            ]):
                raise FileNotFoundError("One or more input files are missing")
            return True
        except Exception as e:
            raise Exception(f"File validation error: {str(e)}")

    @staticmethod
    def get_actual_column_name(expected_name, column_names):
        """Match column names case-insensitively."""
        for col in column_names:
            if col and col.strip().lower() == expected_name.strip().lower():
                return col
        raise KeyError(f"'{expected_name}' column not found in the worksheet.")

    def process(self):
        try:
            logger.info("Starting file processing...")
            self.validate_files()
            
            # Load base Excel file
            logger.info("Loading Excel workbook...")
            wb = load_workbook(self.base_file)
            far_right_sheet = wb.sheetnames[-1]
            ws = wb[far_right_sheet]
            
            # Load the sheet into DataFrame
            data = pd.DataFrame(ws.values)
            header = data.iloc[0]
            data = data[1:]
            data.columns = header

            # Get column names dynamically
            gbil_available_col = self.get_actual_column_name("GBIL Available", data.columns)
            cash_in_account_col = self.get_actual_column_name("Cash in Account", data.columns)
            account_col = self.get_actual_column_name("Account", data.columns)

            logger.info("Processing GBIL data...")
            # Clean account numbers in base data
            data[account_col] = (
                data[account_col]
                .astype(str)
                .apply(lambda x: re.sub(r'\D', '', x))
                .str.lstrip('0')
            )

            # Process GBIL data
            logger.info("Processing GBIL data...")
            gbil_data = pd.read_csv(self.gbil_file)
            
            # Debug: Print all column names
            logger.info(f"Available columns in GBIL file: {gbil_data.columns.tolist()}")
            
            gbil_data['Acct Code'] = (
                gbil_data['Acct Code']
                .astype(str)
                .apply(lambda x: re.sub(r'\D', '', x))
                .str.lstrip('0')
            )

            # Merge GBIL data
            try:
                merged_data = data.merge(
                    gbil_data[['Acct Code', 'Asset Value']],
                    left_on=account_col,
                    right_on='Acct Code',
                    how='left'
                )
                logger.info(f"Columns after merge: {merged_data.columns.tolist()}")
                
                # Use Asset Value_x which is what pandas renamed it to
                merged_data[gbil_available_col] = merged_data['Asset Value_x'].fillna("#N/A")
                
            except KeyError as e:
                logger.error(f"Column error during merge. Available columns: {gbil_data.columns.tolist()}")
                raise
            except IndexError:
                logger.error(f"Asset Value column not found after merge. Available columns: {merged_data.columns.tolist()}")
                raise Exception("Asset Value column not found after merge")

            logger.info("Processing Cash data...")
            # Process Cash data
            logger.info("Processing Cash data...")
            cash_data = pd.read_csv(self.cash_file)
            logger.info(f"Cash CSV columns: {cash_data.columns.tolist()}")
            
            cash_data['Account Number'] = (
                cash_data['Account Number']
                .astype(str)
                .apply(lambda x: re.sub(r'\D', '', x))
                .str.lstrip('0')
            )

            # Merge Cash data
            merged_data = merged_data.merge(
                cash_data[['Account Number', 'Cash Value']],
                left_on=account_col,
                right_on='Account Number',
                how='left'
            )
            logger.info(f"Columns after cash merge: {merged_data.columns.tolist()}")

            # Use the renamed column after merge
            merged_data[cash_in_account_col] = merged_data['Cash Value_x'].fillna("#N/A")

            logger.info("Updating worksheet...")
            # Update worksheet
            columns_to_update = [gbil_available_col, cash_in_account_col]
            for col in columns_to_update:
                col_index = list(header).index(col) + 1
                for row_idx, value in enumerate(merged_data[col], start=2):
                    ws.cell(row=row_idx, column=col_index, value=value)

            # Update date
            ws["L2"].value = datetime.now().strftime("%m/%d/%Y")

            # Save workbook
            logger.info("Saving workbook...")
            wb.save(self.base_file)
            logger.info("Processing completed successfully.")

        except Exception as e:
            logger.error(f"Processing failed: {e}")
            raise

if __name__ == "__main__":
    try:
        # Define file paths
        base_file = r"\\HWM29-LT\Users\ErikKnudsen\Hohimer Wealth Management\Investment Team - Alternative Investments\Structured Notes Allocations - TEST.xlsx"
        gbil_file = r"\\HWM29-LT\Users\ErikKnudsen\Hohimer Wealth Management\Investment Team - Alternative Investments\GBIL and Cash - Raw Data CSV\GBILL\Account Search by Ticker and As Of Date (Greater than Zero Assets Only)_HohimerWea.csv"
        cash_file = r"\\HWM29-LT\Users\ErikKnudsen\Hohimer Wealth Management\Investment Team - Alternative Investments\GBIL and Cash - Raw Data CSV\CASH\Hohimer Wealth - Cash Percentage in Registration_HohimerWea.csv"

        # Initialize and run processor
        processor = StructuredNotesProcessor(base_file, gbil_file, cash_file)
        processor.process()
        print("Processing completed successfully!")
        
    except Exception as e:
        print(f"Error: {e}")
        logger.error(f"Main execution failed: {e}", exc_info=True)
