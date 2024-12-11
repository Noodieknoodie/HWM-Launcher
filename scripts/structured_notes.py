from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QPushButton, QLabel, 
                           QFileDialog, QMessageBox, QGridLayout, QFrame,
                           QHBoxLayout, QProgressDialog)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import pandas as pd
import re
from datetime import datetime
import os
from utils.theme import Theme

class StructuredNotesProcessor:
    def __init__(self):
        self.original_file = None
        self.gbil_file = None
        self.cash_file = None
        self.last_output_path = None
        
    def process_files(self, output_path, progress_callback=None):
        try:
            if progress_callback:
                progress_callback(10, "Loading Excel workbook...")
                
            # Load the original workbook and far-right worksheet
            wb = load_workbook(self.original_file)
            far_right_sheet = wb.sheetnames[-1]
            ws = wb[far_right_sheet]

            # Load the far-right sheet into a DataFrame
            data = pd.DataFrame(ws.values)
            header = data.iloc[0]
            data = data[1:]
            data.columns = header

            if progress_callback:
                progress_callback(30, "Processing GBIL data...")
                
            # Process GBIL data
            first_csv = pd.read_csv(self.gbil_file)
            first_csv['Acct Code'] = (
                first_csv['Acct Code']
                .astype(str)
                .apply(lambda x: re.sub(r'\D', '', x))
                .str.lstrip('0')
            )

            # Get column names dynamically
            gbil_available_col = self.get_actual_column_name("GBIL Available", data.columns)
            cash_in_account_col = self.get_actual_column_name("Cash in Account", data.columns)
            account_col = self.get_actual_column_name("Account", data.columns)

            data[account_col] = (
                data[account_col]
                .astype(str)
                .apply(lambda x: re.sub(r'\D', '', x))
                .str.lstrip('0')
            )

            # Merge and process data
            merged_data = data.merge(
                first_csv[['Acct Code', 'Asset Value']],
                left_on=account_col,
                right_on='Acct Code',
                how='left'
            )
            
            merged_data[gbil_available_col] = merged_data['Asset Value'].fillna("#N/A")

            if progress_callback:
                progress_callback(60, "Processing Cash data...")
                
            # Process Cash data
            second_csv = pd.read_csv(self.cash_file)
            second_csv['Account Number'] = (
                second_csv['Account Number']
                .astype(str)
                .apply(lambda x: re.sub(r'\D', '', x))
                .str.lstrip('0')
            )

            merged_data = merged_data.merge(
                second_csv[['Account Number', 'Cash Value']],
                left_on=account_col,
                right_on='Account Number',
                how='left'
            )

            merged_data[cash_in_account_col] = merged_data['Cash Value'].fillna("#N/A")

            if progress_callback:
                progress_callback(90, "Saving updated workbook...")
                
            # Update worksheet
            columns_to_update = [gbil_available_col, cash_in_account_col]
            for col in columns_to_update:
                col_index = list(header).index(col) + 1
                for row_idx, value in enumerate(merged_data[col], start=2):
                    ws.cell(row=row_idx, column=col_index, value=value)

            # Update date
            ws["L2"].value = datetime.now().strftime("%m/%d/%Y")

            # Save workbook
            wb.save(output_path)
            
            if progress_callback:
                progress_callback(100, "Complete!")
                
            return True

        except Exception as e:
            raise Exception(f"Processing error: {str(e)}")

    @staticmethod
    def get_actual_column_name(expected_name, column_names):
        for col in column_names:
            if col and col.strip().lower() == expected_name.strip().lower():
                return col
        raise KeyError(f"'{expected_name}' column not found in the worksheet.")

    def validate_files(self):
        """Validate all input files before processing."""
        try:
            if not all([
                os.path.exists(self.original_file),
                os.path.exists(self.gbil_file),
                os.path.exists(self.cash_file)
            ]):
                raise FileNotFoundError("One or more input files are missing")
            return True
        except Exception as e:
            raise Exception(f"File validation error: {str(e)}")

def load_ui(parent=None):
    """Create and return the UI for the structured notes processor."""
    widget = QWidget(parent)
    main_layout = QVBoxLayout(widget)
    main_layout.setContentsMargins(Theme.PADDING['large'], 
                                 Theme.PADDING['large'], 
                                 Theme.PADDING['large'], 
                                 Theme.PADDING['large'])
    
    # Create processor instance
    processor = StructuredNotesProcessor()
    
    # Header Section
    header_frame = QFrame()
    header_frame.setObjectName("header_frame")
    header_layout = QVBoxLayout(header_frame)
    
    title_label = QLabel("Structured Notes Processor")
    title_label.setFont(Theme.get_title_font())
    title_label.setAlignment(Qt.AlignCenter)
    
    description = QLabel(
        "Upload the required files in sequence to process structured notes data.\n"
        "All files must be uploaded before processing can begin."
    )
    description.setAlignment(Qt.AlignCenter)
    description.setFont(Theme.get_body_font())
    description.setStyleSheet(f"color: {Theme.TEXT_SECONDARY};")
    
    header_layout.addWidget(title_label)
    header_layout.addWidget(description)
    main_layout.addWidget(header_frame)
    
    # Separator
    separator = QFrame()
    separator.setFrameShape(QFrame.HLine)
    separator.setFrameShadow(QFrame.Sunken)
    separator.setStyleSheet(f"background-color: {Theme.BORDER};")
    main_layout.addWidget(separator)
    
    # Files Section
    files_frame = QFrame()
    files_frame.setObjectName("files_frame")
    files_layout = QVBoxLayout(files_frame)
    
    def create_file_section(title, button_text, icon_name=None):
        section = QFrame()
        section.setObjectName("file_section")
        section_layout = QHBoxLayout(section)
        
        # Info Column
        info_layout = QVBoxLayout()
        title_label = QLabel(title)
        title_label.setObjectName("section_title")
        
        status_label = QLabel("No file selected")
        status_label.setObjectName("status_label")
        status_label.setWordWrap(True)
        
        info_layout.addWidget(title_label)
        info_layout.addWidget(status_label)
        section_layout.addLayout(info_layout, stretch=2)
        
        # Button Column
        button = QPushButton(button_text)
        button.setStyleSheet(Theme.BUTTON_STYLE.format(**Theme.get_style_params()))
        section_layout.addWidget(button, stretch=1, alignment=Qt.AlignVCenter)
        
        section.setStyleSheet(Theme.FILE_SECTION_STYLE.format(**Theme.get_style_params()))
        
        return section, button, status_label
    
    # Create file sections
    excel_section, excel_btn, excel_status = create_file_section(
        "Excel File", "Upload Excel", "excel-icon")
    gbil_section, gbil_btn, gbil_status = create_file_section(
        "GBIL CSV", "Upload GBIL", "csv-icon")
    cash_section, cash_btn, cash_status = create_file_section(
        "Cash CSV", "Upload Cash", "csv-icon")
    
    files_layout.addWidget(excel_section)
    files_layout.addWidget(gbil_section)
    files_layout.addWidget(cash_section)
    main_layout.addWidget(files_frame)
    
    # Action Section
    action_frame = QFrame()
    action_frame.setObjectName("action_frame")
    action_layout = QVBoxLayout(action_frame)
    
    process_btn = QPushButton("UPDATE")
    process_btn.setEnabled(False)
    process_btn.setMinimumHeight(50)
    process_btn.setStyleSheet(Theme.BUTTON_STYLE.format(**Theme.get_style_params()))
    
    status_label = QLabel("Please upload all required files")
    status_label.setAlignment(Qt.AlignCenter)
    status_label.setStyleSheet(f"color: {Theme.TEXT_SECONDARY}; font-style: italic;")
    
    action_layout.addWidget(process_btn)
    action_layout.addWidget(status_label)
    main_layout.addWidget(action_frame)
    
    # Add stretcher to keep content at top
    main_layout.addStretch()

    # Add event handlers for buttons
    def update_status():
        """Update status label and process button state"""
        files_loaded = all([
            processor.original_file,
            processor.gbil_file,
            processor.cash_file
        ])
        process_btn.setEnabled(files_loaded)
        if files_loaded:
            status_label.setText("Ready to process!")
            status_label.setStyleSheet(f"color: {Theme.SUCCESS}; font-style: normal; font-weight: bold;")
        else:
            status_label.setText("Please upload all required files")
            status_label.setStyleSheet(f"color: {Theme.TEXT_SECONDARY}; font-style: italic;")
    
    def load_excel():
        file_path, _ = QFileDialog.getOpenFileName(
            widget,
            "Select Structured Notes Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            try:
                # Validate Excel file
                wb = load_workbook(file_path, read_only=True)
                wb.close()
                
                processor.original_file = file_path
                excel_status.setText(f"Selected: {os.path.basename(file_path)}")
                excel_status.setStyleSheet(f"color: {Theme.SUCCESS};")
                update_status()
            except Exception as e:
                QMessageBox.warning(
                    widget,
                    "Invalid File",
                    f"The selected Excel file appears to be invalid:\n{str(e)}"
                )
    
    def load_gbil():
        file_path, _ = QFileDialog.getOpenFileName(
            widget,
            "Select GBIL CSV File",
            "",
            "CSV Files (*.csv)"
        )
        if file_path:
            try:
                # Validate CSV file
                pd.read_csv(file_path, nrows=1)
                
                processor.gbil_file = file_path
                gbil_status.setText(f"Selected: {os.path.basename(file_path)}")
                gbil_status.setStyleSheet(f"color: {Theme.SUCCESS};")
                update_status()
            except Exception as e:
                QMessageBox.warning(
                    widget,
                    "Invalid File",
                    f"The selected GBIL CSV file appears to be invalid:\n{str(e)}"
                )
    
    def load_cash():
        file_path, _ = QFileDialog.getOpenFileName(
            widget,
            "Select Cash CSV File",
            "",
            "CSV Files (*.csv)"
        )
        if file_path:
            try:
                # Validate CSV file
                pd.read_csv(file_path, nrows=1)
                
                processor.cash_file = file_path
                cash_status.setText(f"Selected: {os.path.basename(file_path)}")
                cash_status.setStyleSheet(f"color: {Theme.SUCCESS};")
                update_status()
            except Exception as e:
                QMessageBox.warning(
                    widget,
                    "Invalid File",
                    f"The selected Cash CSV file appears to be invalid:\n{str(e)}"
                )
    
    def process_files():
        try:
            processor.validate_files()
            
            output_path, _ = QFileDialog.getSaveFileName(
                widget,
                "Save Updated Excel File",
                os.path.dirname(processor.original_file),
                "Excel Files (*.xlsx)"
            )
            
            if output_path:
                progress = QProgressDialog("Processing files...", "Cancel", 0, 100, widget)
                progress.setWindowModality(Qt.WindowModal)
                progress.setWindowTitle("Processing")
                progress.setAutoClose(True)
                progress.setAutoReset(True)
                progress.setMinimumDuration(0)
                
                def update_progress(value, message):
                    if progress.wasCanceled():
                        raise Exception("Operation cancelled by user")
                    progress.setValue(value)
                    progress.setLabelText(message)
                
                try:
                    processor.process_files(output_path, progress_callback=update_progress)
                    QMessageBox.information(
                        widget,
                        "Success",
                        "Files processed successfully!"
                    )
                except Exception as e:
                    raise e
                finally:
                    progress.close()
                    
        except Exception as e:
            if str(e) == "Operation cancelled by user":
                QMessageBox.information(
                    widget,
                    "Cancelled",
                    "Operation was cancelled by user."
                )
            else:
                QMessageBox.critical(
                    widget,
                    "Error",
                    f"An error occurred while processing the files:\n{str(e)}"
                )
    
    # Connect buttons to functions
    excel_btn.clicked.connect(load_excel)
    gbil_btn.clicked.connect(load_gbil)
    cash_btn.clicked.connect(load_cash)
    process_btn.clicked.connect(process_files)
    
    return widget 